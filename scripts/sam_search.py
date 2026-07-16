#!/usr/bin/env python3
"""
SAM.gov daily opportunity search.

- Spaces out API requests (gentle defaults)
- Accumulates hits into a day-by-day history store
- Writes Excel + HTML trackers (latest + dated copies)

Get Opportunities Public API:
  https://open.gsa.gov/api/get-opportunities-public-api/
"""

from __future__ import annotations

import argparse
import html
import json
import os
import random
import re
import shutil
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any

_TAG_RE = re.compile(r"<[^>]+>")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover
    print(
        "Missing openpyxl. From the tool root run:\n"
        "  python3 -m venv .venv && .venv/bin/pip install -r requirements.txt\n"
        "  .venv/bin/python scripts/sam_search.py ...",
        file=sys.stderr,
    )
    raise SystemExit(3)

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_TERMS = ROOT / "config" / "search_terms.json"
DEFAULT_HISTORY = ROOT / "data" / "history.json"
DEFAULT_API_ROTATION = ROOT / "data" / "api_term_rotation.json"
DEFAULT_REPORTS = ROOT / "reports"
DEFAULT_BASE_URL = "https://api.sam.gov/prod/opportunities/v2/search"
# Same search backend the SAM.gov website UI uses (no public API key quota).
DEFAULT_FRONTEND_URL = "https://sam.gov/api/prod/sgs/v1/search/"
# Optional local folder for latest xlsx/html (set SAM_LATEST_SYNC_DIR in .env).
# No personal paths in the public repo defaults.
DEFAULT_LATEST_SYNC_DIR: Path | None = None

# Gentle defaults — do not hammer SAM
# API source uses public Opportunities key (personal keys often ~10 req/day).
# We hit 429 after ~7 successful calls; default batch 7 ⇒ full 22-term cycle in 4 days.
DEFAULT_API_BATCH_SIZE = 7
DEFAULT_TERM_SLEEP_API = 5.0
DEFAULT_PAGE_SLEEP_API = 2.0
# Frontend: deliberately human-slow (not fixed robot intervals).
# Typical gap between keyword searches ~20–50s; sometimes longer "reading" pauses.
DEFAULT_TERM_SLEEP_FRONTEND = 28.0  # center of human gap (seconds)
DEFAULT_PAGE_SLEEP_FRONTEND = 8.0
DEFAULT_JITTER = 1.0  # only used for --source api; frontend uses human_pause()
DEFAULT_MAX_PAGES = 1
DEFAULT_LIMIT = 25
# Keep day buckets / notices / run log for this many calendar days (rolling).
DEFAULT_HISTORY_RETENTION_DAYS = 15
# Default posted-date look-back when searching (also fills a 15-day record window).
DEFAULT_SEARCH_DAYS = 15

# A few realistic browser UA strings (rotated randomly — not cycled on a timer)
BROWSER_USER_AGENTS = [
    (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    ),
    (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 "
        "(KHTML, like Gecko) Version/17.2 Safari/605.1.15"
    ),
    (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_2) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36"
    ),
]

PTYPE_LABELS = {
    "u": "Justification (J&A)",
    "p": "Presolicitation",
    "a": "Award Notice",
    "r": "Sources Sought",
    "s": "Special Notice",
    "o": "Solicitation",
    "g": "Sale of Surplus Property",
    "k": "Combined Synopsis/Solicitation",
    "i": "Intent to Bundle Requirements (DoD-Funded)",
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
CELL_FONT = Font(name="Arial", size=10)
LINK_FONT = Font(name="Arial", size=10, color="0563C1", underline="single")
NEW_FILL = PatternFill("solid", fgColor="E2EFDA")
DAY_FILL = PatternFill("solid", fgColor="D6EAF8")
THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


@dataclass
class Hit:
    notice_id: str
    title: str
    solicitation_number: str
    posted_date: str
    type: str
    response_deadline: str
    naics: str
    active: str
    ui_link: str
    organization: str
    matched_terms: set[str] = field(default_factory=set)
    award_amount: str = ""
    awardee: str = ""

    def public_url(self) -> str:
        if self.ui_link and self.ui_link not in ("null", "None"):
            return self.ui_link.replace("https://beta.sam.gov/", "https://sam.gov/")
        if self.notice_id:
            return f"https://sam.gov/opp/{self.notice_id}/view"
        return ""

    def record_key(self) -> str:
        return self.notice_id or f"{self.solicitation_number}|{self.title}|{self.posted_date}"


def normalize_api_key(raw: str) -> str:
    val = (raw or "").strip().strip("'").strip('"')
    for prefix in ("SAM_API_KEY=", "api_key=", "API_KEY=", "apikey="):
        if val.lower().startswith(prefix.lower()):
            val = val[len(prefix) :].strip().strip("'").strip('"')
            break
    return val.strip()


def load_dotenv(path: Path) -> None:
    if not path.is_file():
        return
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, _, value = line.partition("=")
        key = key.strip()
        value = value.strip().strip("'").strip('"')
        if key == "SAM_API_KEY":
            value = normalize_api_key(value)
        if key and key not in os.environ:
            os.environ[key] = value


def load_api_key_from_files() -> str | None:
    candidates = [ROOT / ".env", ROOT / "api.txt", ROOT.parent / "api.txt"]
    for path in candidates:
        if not path.is_file():
            continue
        text = path.read_text(encoding="utf-8").strip()
        if path.name == ".env":
            for line in text.splitlines():
                if line.strip().startswith("SAM_API_KEY="):
                    return normalize_api_key(line.split("=", 1)[1])
            continue
        if text:
            return normalize_api_key(text.splitlines()[0])
    return None


def load_terms(path: Path) -> list[str]:
    data = json.loads(path.read_text(encoding="utf-8"))
    terms = data.get("terms") if isinstance(data, dict) else data
    if not isinstance(terms, list) or not terms:
        raise SystemExit(f"No terms found in {path}")
    seen: set[str] = set()
    out: list[str] = []
    for t in terms:
        t = str(t).strip()
        if not t:
            continue
        k = t.casefold()
        if k in seen:
            continue
        seen.add(k)
        out.append(t)
    return out


def fmt_mmddyyyy(d: date) -> str:
    return d.strftime("%m/%d/%Y")


def parse_date(s: str) -> date:
    return datetime.strptime(s, "%Y-%m-%d").date()


def paced_sleep(base: float, jitter: float, *, label: str | None = None) -> None:
    """Simple sleep for API mode (base + uniform jitter)."""
    delay = max(0.0, base) + (random.uniform(0, jitter) if jitter > 0 else 0.0)
    if delay <= 0:
        return
    _sleep_with_countdown(delay, label)


def _sleep_with_countdown(delay: float, label: str | None) -> None:
    if label and sys.stdout.isatty() and delay >= 1.0:
        end = time.monotonic() + delay
        while True:
            left = end - time.monotonic()
            if left <= 0:
                break
            sys.stdout.write(f"\r  ⏳ {label}: ~{left:5.1f}s remaining...   ")
            sys.stdout.flush()
            time.sleep(min(0.35, left))
        sys.stdout.write("\r" + " " * 72 + "\r")
        sys.stdout.flush()
    else:
        time.sleep(delay)


def human_pause(
    *,
    kind: str = "between_terms",
    hits_this: int = 0,
    label: str | None = None,
) -> float:
    """
    Human-like variable pauses (not a fixed robot interval).

    kind:
      - think: short pause before submitting a search (typing / choosing keywords)
      - read: skim results after a search
      - between_terms: idle between different keyword searches
      - between_pages: pause before next result page
    """
    if kind == "think":
        # 1.5–5s thinking / "typing"
        delay = random.triangular(1.5, 5.0, 2.8)
    elif kind == "read":
        # Longer if there were hits to "skim"
        base = random.triangular(4.0, 14.0, 7.0)
        delay = base + min(hits_this, 15) * random.uniform(0.4, 1.1)
    elif kind == "between_pages":
        delay = random.triangular(5.0, 18.0, 9.0)
    else:  # between_terms
        # Most gaps ~18–45s; occasional longer break like a person multitasking
        if random.random() < 0.12:
            delay = random.uniform(55.0, 110.0)  # short break
            label = label or "human break"
        else:
            delay = random.triangular(18.0, 48.0, 28.0)

    # Micro-jitter so we never hit the same second twice in a row
    delay += random.uniform(0.05, 1.25)
    _sleep_with_countdown(delay, label or kind)
    return delay


def _fmt_duration(seconds: float) -> str:
    seconds = max(0, int(seconds))
    m, s = divmod(seconds, 60)
    h, m = divmod(m, 60)
    if h:
        return f"{h}:{m:02d}:{s:02d}"
    return f"{m}:{s:02d}"


def print_progress(
    *,
    done: int,
    total: int,
    term: str,
    hits_this: int | None = None,
    total_hits: int = 0,
    errors: int = 0,
    started: float,
    phase: str = "done",
) -> None:
    """Terminal progress bar + status. Safe for logs (full line when not a TTY)."""
    total = max(total, 1)
    frac = min(1.0, done / total)
    width = 24
    filled = int(width * frac)
    bar = "█" * filled + "░" * (width - filled)
    pct = 100.0 * frac
    elapsed = time.monotonic() - started
    if done > 0:
        eta = (elapsed / done) * (total - done)
        eta_s = _fmt_duration(eta)
    else:
        eta_s = "—"
    hit_part = f"→ {hits_this} hit(s)" if hits_this is not None else "requesting…"
    line = (
        f"[{bar}] {done}/{total} ({pct:5.1f}%)  "
        f"{phase} {term!r} {hit_part}  | total_hits={total_hits} err={errors}  "
        f"elapsed {_fmt_duration(elapsed)}  eta ~{eta_s}"
    )
    if sys.stdout.isatty():
        # In-place bar for the active term; final state ends with newline via caller
        sys.stdout.write("\r" + line[:140].ljust(140))
        sys.stdout.flush()
    else:
        # Log-friendly: one permanent line per update
        print(line, flush=True)


def http_get_json(
    url: str,
    *,
    timeout: float = 60.0,
    headers: dict[str, str] | None = None,
    redact_api_key: bool = True,
) -> dict[str, Any]:
    hdrs = headers or {
        "Accept": "application/json",
        "User-Agent": "sam-daily-search/2.1",
    }
    req = urllib.request.Request(url, headers=hdrs, method="GET")
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            body = resp.read().decode("utf-8")
    except urllib.error.HTTPError as e:
        detail = e.read().decode("utf-8", errors="replace")
        safe = url
        if redact_api_key and "api_key=" in url:
            safe = url.split("api_key=")[0] + "api_key=***"
        raise RuntimeError(f"HTTP {e.code} for {safe}: {detail[:500]}") from e
    except urllib.error.URLError as e:
        raise RuntimeError(f"Network error: {e}") from e
    try:
        return json.loads(body)
    except json.JSONDecodeError as e:
        raise RuntimeError(f"Non-JSON response: {body[:300]}") from e


def opportunity_from_row(row: dict[str, Any], term: str) -> Hit:
    award = row.get("award") or {}
    awardee = ""
    amount = ""
    if isinstance(award, dict):
        amount = str(award.get("amount") or "")
        aw = award.get("awardee") or {}
        if isinstance(aw, dict):
            awardee = str(aw.get("name") or "")
        elif aw:
            awardee = str(aw)

    org = row.get("fullParentPathName") or ""
    if not org:
        parts = [row.get("department"), row.get("subTier"), row.get("office")]
        org = ".".join(p for p in parts if p)

    return Hit(
        notice_id=str(row.get("noticeId") or "").strip(),
        title=str(row.get("title") or "").strip(),
        solicitation_number=str(row.get("solicitationNumber") or "").strip(),
        posted_date=str(row.get("postedDate") or "").strip()[:10],
        type=str(row.get("type") or "").strip(),
        response_deadline=str(row.get("responseDeadLine") or row.get("reponseDeadLine") or "").strip(),
        naics=str(row.get("naicsCode") or "").strip(),
        active=str(row.get("active") or "").strip(),
        ui_link=str(row.get("uiLink") or "").strip(),
        organization=str(org).strip(),
        matched_terms={term},
        award_amount=amount,
        awardee=awardee,
    )


def search_term_api(
    *,
    base_url: str,
    api_key: str,
    term: str,
    posted_from: date,
    posted_to: date,
    ptypes: list[str] | None,
    limit: int,
    max_pages: int,
    page_sleep: float,
    jitter: float,
) -> list[Hit]:
    """Public Opportunities API (requires SAM public API key; daily quota)."""
    hits: list[Hit] = []
    offset = 0
    pages = 0
    total_records: int | None = None

    while pages < max_pages:
        params: dict[str, str | int] = {
            "api_key": api_key,
            "postedFrom": fmt_mmddyyyy(posted_from),
            "postedTo": fmt_mmddyyyy(posted_to),
            # Multi-word title phrases quoted when possible
            "title": search_query_for_term(term).strip('"'),
            "limit": limit,
            "offset": offset,
        }
        query_parts = [urllib.parse.urlencode(params)]
        if ptypes:
            for pt in ptypes:
                query_parts.append(urllib.parse.urlencode({"ptype": pt}))
        url = f"{base_url}?{'&'.join(query_parts)}"

        data = http_get_json(url)
        if total_records is None:
            total_records = int(data.get("totalRecords") or 0)

        rows = data.get("opportunitiesData") or []
        if not isinstance(rows, list):
            break
        for row in rows:
            if isinstance(row, dict):
                hits.append(opportunity_from_row(row, term))

        pages += 1
        offset += limit
        if not rows or offset >= (total_records or 0):
            break
        paced_sleep(page_sleep, jitter)

    return hits


def _parse_iso_date(val: Any) -> date | None:
    if not val:
        return None
    s = str(val).strip()
    if not s:
        return None
    # 2026-07-15T19:43:27+00:00 or 2026-07-15
    try:
        return date.fromisoformat(s[:10])
    except ValueError:
        return None


def is_multi_word_term(term: str) -> bool:
    return len(str(term).strip().split()) >= 2


def search_query_for_term(term: str) -> str:
    """
    Multi-word terms are wrapped in double quotes for phrase search.
    Single-word terms are left as-is. Existing quotes are preserved.
    """
    t = str(term).strip()
    if not t:
        return t
    if (t.startswith('"') and t.endswith('"')) or (t.startswith("'") and t.endswith("'")):
        return t if t.startswith('"') else f'"{t[1:-1]}"'
    if is_multi_word_term(t):
        return f'"{t}"'
    return t


def _strip_html(text: str) -> str:
    return _TAG_RE.sub(" ", text or "")


def _row_text_blob(row: dict[str, Any], title: str) -> str:
    parts = [title or ""]
    descs = row.get("descriptions") or []
    if isinstance(descs, list):
        for d in descs:
            if isinstance(d, dict):
                parts.append(_strip_html(str(d.get("content") or "")))
            elif d:
                parts.append(_strip_html(str(d)))
    return " ".join(parts)


def term_matches_result(term: str, title: str, blob: str = "") -> bool:
    """
    Single-word: accept (SGS already filtered).
    Multi-word: require the full phrase in title or description (case-insensitive).
    This blocks false hits on isolated words like 'Digital' or 'Laboratory'.
    """
    t = str(term).strip().strip('"').strip("'")
    if not is_multi_word_term(t):
        return True
    phrase = t.casefold()
    hay_title = (title or "").casefold()
    if phrase in hay_title:
        return True
    hay = (blob or title or "").casefold()
    return phrase in hay


def opportunity_from_frontend_row(row: dict[str, Any], term: str) -> Hit | None:
    """Map SAM website SGS search result → Hit. None if multi-word phrase not present."""
    notice_id = str(row.get("_id") or row.get("parentNoticeId") or "").strip()
    type_obj = row.get("type") or {}
    type_s = ""
    if isinstance(type_obj, dict):
        type_s = str(type_obj.get("value") or type_obj.get("code") or "")
    else:
        type_s = str(type_obj)

    org_parts: list[str] = []
    oh = row.get("organizationHierarchy") or []
    if isinstance(oh, list):
        for level in oh:
            if isinstance(level, dict) and level.get("name"):
                org_parts.append(str(level["name"]))
    org = ".".join(org_parts)

    award = row.get("award") or {}
    awardee = ""
    amount = ""
    if isinstance(award, dict):
        amount = str(award.get("amount") or "")
        aw = award.get("awardee") or {}
        if isinstance(aw, dict):
            awardee = str(aw.get("name") or "")

    pub = _parse_iso_date(row.get("publishDate"))
    resp = row.get("responseDate") or row.get("responseDateActual") or ""
    active = "Yes" if row.get("isActive") else "No"
    title = str(row.get("title") or "").strip()
    blob = _row_text_blob(row, title)
    if not term_matches_result(term, title, blob):
        return None

    return Hit(
        notice_id=notice_id,
        title=title,
        solicitation_number=str(row.get("solicitationNumber") or "").strip(),
        posted_date=pub.isoformat() if pub else str(row.get("publishDate") or "")[:10],
        type=type_s,
        response_deadline=str(resp).strip(),
        naics="",  # not always present on SGS list cards
        active=active,
        ui_link=f"https://sam.gov/opp/{notice_id}/view" if notice_id else "",
        organization=org,
        matched_terms={term},
        award_amount=amount,
        awardee=awardee,
    )


def _frontend_headers(term: str) -> dict[str, str]:
    """Browser-like headers; Referer matches a real sam.gov search URL for the term."""
    q_display = search_query_for_term(term)
    q = urllib.parse.quote(q_display)
    referer = (
        "https://sam.gov/search/?index=opp&page=1&pageSize=25&sort=-modifiedDate"
        f"&sfm%5BsimpleSearch%5D%5BkeywordRadio%5D=ALL"
        f"&sfm%5BsimpleSearch%5D%5BkeywordTags%5D%5B0%5D%5Bkey%5D={q}"
        f"&sfm%5BsimpleSearch%5D%5BkeywordTags%5D%5B0%5D%5Bvalue%5D={q}"
    )
    return {
        "Accept": "application/hal+json",
        "User-Agent": random.choice(BROWSER_USER_AGENTS),
        "Referer": referer,
        "Origin": "https://sam.gov",
        "Accept-Language": "en-US,en;q=0.9",
        "Accept-Encoding": "gzip, deflate, br",
        "Connection": "keep-alive",
        "Sec-Fetch-Dest": "empty",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Site": "same-origin",
    }


def search_term_frontend(
    *,
    base_url: str,
    term: str,
    posted_from: date,
    posted_to: date,
    limit: int,
    max_pages: int,
    active_only: bool,
    human: bool = True,
) -> list[Hit]:
    """
    Same search backend the SAM.gov website UI uses (SGS).
    No public API key — not the Opportunities API quota.
    When human=True: variable pauses between pages (no fixed robot cadence).
    """
    hits: list[Hit] = []
    page = 0
    total_pages: int | None = None

    # Humans sometimes change page size slightly
    page_size = limit
    if human:
        page_size = int(random.choice([20, 25, 25, 25, 50]))
        page_size = min(max(page_size, 1), 100)

    while page < max_pages:
        if human and page == 0:
            human_pause(kind="think", label=f"before search {term!r}")
        elif human and page > 0:
            human_pause(kind="between_pages", label=f"next page of {term!r}")

        q = search_query_for_term(term)
        headers = _frontend_headers(term)
        params: list[tuple[str, str]] = [
            ("index", "opp"),
            ("page", str(page)),
            ("size", str(page_size)),
            ("mode", "search"),
            ("sort", "-modifiedDate"),
            # Multi-word → "quoted phrase"; single-word unchanged
            ("q", q),
        ]
        if active_only:
            params.append(("is_active", "true"))
        url = base_url.rstrip("/") + "/?" + urllib.parse.urlencode(params)

        # urllib does not always decode br/gzip without extra care; prefer identity if needed
        headers_req = dict(headers)
        headers_req["Accept-Encoding"] = "identity"

        data = http_get_json(url, headers=headers_req, redact_api_key=False, timeout=45.0)
        page_info = data.get("page") or {}
        if total_pages is None:
            total_pages = int(page_info.get("totalPages") or 1)

        rows = (data.get("_embedded") or {}).get("results") or []
        if not isinstance(rows, list):
            break

        page_hits = 0
        for row in rows:
            if not isinstance(row, dict):
                continue
            h = opportunity_from_frontend_row(row, term)
            if h is None:
                continue  # multi-word phrase not in title/body
            pd = _parse_iso_date(h.posted_date) or _parse_iso_date(row.get("publishDate"))
            if pd is not None and (pd < posted_from or pd > posted_to):
                continue
            hits.append(h)
            page_hits += 1

        if human:
            human_pause(kind="read", hits_this=page_hits, label=f"reading {term!r} results")

        page += 1
        if page >= (total_pages or 1) or not rows:
            break

    return hits


def merge_hits(all_hits: list[Hit]) -> list[Hit]:
    by_id: dict[str, Hit] = {}
    for h in all_hits:
        key = h.record_key()
        if key in by_id:
            by_id[key].matched_terms |= h.matched_terms
        else:
            by_id[key] = h
    # Drop multi-word false matches that only survived via OR-style SGS ranking
    cleaned: list[Hit] = []
    for h in by_id.values():
        kept = {t for t in h.matched_terms if term_matches_result(t, h.title, h.title)}
        if not kept:
            continue
        h.matched_terms = kept
        cleaned.append(h)
    return sorted(cleaned, key=lambda x: (x.posted_date or "", x.title or ""), reverse=True)


def filter_notice_row(row: dict[str, Any]) -> dict[str, Any] | None:
    """Re-apply multi-word phrase rules to stored history rows (for report rebuilds)."""
    title = str(row.get("title") or "")
    terms = row.get("matched_terms") or []
    if isinstance(terms, str):
        terms = [t.strip() for t in terms.split(";") if t.strip()]
    kept = [t for t in terms if term_matches_result(str(t), title, title)]
    if not kept:
        return None
    out = dict(row)
    out["matched_terms"] = kept
    return out


# --- History store (day-by-day) -------------------------------------------------


def empty_history() -> dict[str, Any]:
    return {"version": 2, "updated_at": None, "runs": [], "notices": {}, "days": {}}


def load_history(path: Path) -> dict[str, Any]:
    if not path.is_file():
        return empty_history()
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return empty_history()
    if not isinstance(data, dict):
        return empty_history()
    data.setdefault("version", 2)
    data.setdefault("runs", [])
    data.setdefault("notices", {})
    data.setdefault("days", {})
    return data


def prune_history(
    history: dict[str, Any],
    *,
    keep_days: int = DEFAULT_HISTORY_RETENTION_DAYS,
    as_of: date | None = None,
) -> dict[str, Any]:
    """
    Keep only the last `keep_days` calendar days of:
      - days[] buckets (report sections)
      - runs[] log entries
      - notices last seen within the window (or first seen if no last)
    """
    if keep_days < 1:
        return history
    today = as_of or date.today()
    cutoff = today - timedelta(days=keep_days - 1)
    cutoff_s = cutoff.isoformat()

    days = history.get("days") or {}
    history["days"] = {
        d: bucket
        for d, bucket in days.items()
        if isinstance(d, str) and d >= cutoff_s
    }

    runs = history.get("runs") or []
    history["runs"] = [
        r
        for r in runs
        if isinstance(r, dict) and str(r.get("run_date") or "") >= cutoff_s
    ]

    notices = history.get("notices") or {}
    kept_notices: dict[str, Any] = {}
    for key, rec in notices.items():
        if not isinstance(rec, dict):
            continue
        last = str(rec.get("last_seen_date") or rec.get("first_seen_date") or "")
        posted = str(rec.get("posted_date") or "")[:10]
        # Keep if we saw it in-window, or it was posted in-window
        if (last and last >= cutoff_s) or (posted and posted >= cutoff_s):
            # Trim seen_on_dates to window
            seen = [d for d in (rec.get("seen_on_dates") or []) if str(d) >= cutoff_s]
            if seen:
                rec = dict(rec)
                rec["seen_on_dates"] = sorted(seen)
            kept_notices[key] = rec
    history["notices"] = kept_notices
    history["retention_days"] = keep_days
    history["retention_cutoff"] = cutoff_s
    return history


def save_history(
    path: Path,
    history: dict[str, Any],
    *,
    keep_days: int = DEFAULT_HISTORY_RETENTION_DAYS,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    prune_history(history, keep_days=max(1, keep_days))
    history["updated_at"] = datetime.now(timezone.utc).isoformat()
    path.write_text(json.dumps(history, indent=2, sort_keys=False) + "\n", encoding="utf-8")


def hit_to_notice_dict(h: Hit) -> dict[str, Any]:
    return {
        "notice_id": h.notice_id,
        "title": h.title,
        "solicitation_number": h.solicitation_number,
        "posted_date": h.posted_date,
        "type": h.type,
        "response_deadline": h.response_deadline,
        "naics": h.naics,
        "active": h.active,
        "organization": h.organization,
        "matched_terms": sorted(h.matched_terms, key=str.casefold),
        "award_amount": h.award_amount,
        "awardee": h.awardee,
        "url": h.public_url(),
    }


def merge_into_history(
    history: dict[str, Any],
    hits: list[Hit],
    *,
    run_date: str,
    run_id: str,
    meta: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Return (day_rows with is_new flag, newly_first_seen rows)."""
    notices: dict[str, Any] = history["notices"]
    days: dict[str, Any] = history["days"]
    day_bucket = days.setdefault(
        run_date,
        {"date": run_date, "runs": [], "notice_keys": [], "new_notice_keys": []},
    )

    day_rows: list[dict[str, Any]] = []
    newly: list[dict[str, Any]] = []

    for h in hits:
        key = h.record_key()
        existing = notices.get(key)
        is_new = existing is None
        if existing is None:
            rec = hit_to_notice_dict(h)
            rec["first_seen_date"] = run_date
            rec["last_seen_date"] = run_date
            rec["seen_on_dates"] = [run_date]
            notices[key] = rec
            newly.append(rec)
            if key not in day_bucket["new_notice_keys"]:
                day_bucket["new_notice_keys"].append(key)
        else:
            # Merge terms + refresh fields
            terms = set(existing.get("matched_terms") or []) | h.matched_terms
            existing["matched_terms"] = sorted(terms, key=str.casefold)
            existing["last_seen_date"] = run_date
            seen_dates = list(existing.get("seen_on_dates") or [])
            if run_date not in seen_dates:
                seen_dates.append(run_date)
            existing["seen_on_dates"] = sorted(seen_dates)
            # Prefer non-empty newer metadata
            for field_name in (
                "title",
                "type",
                "response_deadline",
                "naics",
                "active",
                "organization",
                "award_amount",
                "awardee",
                "url",
                "posted_date",
                "solicitation_number",
            ):
                val = getattr(h, field_name, None) if field_name != "url" else h.public_url()
                if field_name == "url":
                    val = h.public_url()
                elif field_name == "matched_terms":
                    continue
                else:
                    val = getattr(h, field_name)
                if val:
                    existing[field_name] = val
            rec = existing

        if key not in day_bucket["notice_keys"]:
            day_bucket["notice_keys"].append(key)

        row = dict(rec)
        row["tracked_date"] = run_date
        row["is_new"] = is_new
        day_rows.append(row)

    if run_id not in day_bucket["runs"]:
        day_bucket["runs"].append(run_id)

    history["runs"].append(
        {
            "run_id": run_id,
            "run_date": run_date,
            "posted_from": meta.get("posted_from"),
            "posted_to": meta.get("posted_to"),
            "term_count": meta.get("term_count"),
            "hit_count": len(hits),
            "new_count": len(newly),
            "errors": meta.get("errors") or [],
            "term_sleep": meta.get("term_sleep"),
            "page_sleep": meta.get("page_sleep"),
        }
    )

    # New first, then posted_date descending
    day_rows.sort(key=lambda r: (0 if r.get("is_new") else 1, -_date_sort_key(r.get("posted_date"))))
    return day_rows, newly


def _date_sort_key(s: Any) -> int:
    try:
        return int(str(s).replace("-", "")[:8] or "0")
    except ValueError:
        return 0


# --- Excel / HTML outputs -------------------------------------------------------


DETAIL_COLUMNS = [
    ("tracked_date", "Tracked date"),  # A
    ("is_new", "New?"),  # B
    ("posted_date", "Posted"),  # C
    ("title", "Title"),  # D
    ("url", "URL"),  # E — next to title for quick open
    ("matched_terms", "Matched terms"),
    ("type", "Type"),
    ("response_deadline", "Response deadline"),
    ("solicitation_number", "Solicitation #"),
    ("naics", "NAICS"),
    ("organization", "Organization"),
    ("award_amount", "Award amount"),
    ("awardee", "Awardee"),
    ("active", "Active"),
    ("notice_id", "Notice ID"),
    ("first_seen_date", "First seen"),
    ("last_seen_date", "Last seen"),
]


def _style_header(ws, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(1, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN


def _autosize(ws, max_width: int = 48) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = 10
        for cell in col:
            if cell.value is None:
                continue
            width = max(width, min(max_width, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


def _write_detail_rows(ws, rows: list[dict[str, Any]], start_row: int = 2) -> None:
    for r_i, row in enumerate(rows):
        excel_row = start_row + r_i
        for c_i, (key, _) in enumerate(DETAIL_COLUMNS, 1):
            val = row.get(key, "")
            if key == "matched_terms" and isinstance(val, list):
                val = "; ".join(val)
            if key == "is_new":
                val = "YES" if val else ""
            cell = ws.cell(excel_row, c_i, val if val is not None else "")
            cell.border = THIN
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if key == "url" and val and str(val).startswith(("http://", "https://")):
                # Clickable hyperlink (Excel / Numbers / LibreOffice)
                cell.hyperlink = str(val)
                cell.font = LINK_FONT
            else:
                cell.font = CELL_FONT
            if row.get("is_new") and key == "is_new":
                cell.fill = NEW_FILL
        if row.get("is_new"):
            ws.cell(excel_row, 1).fill = NEW_FILL


def write_excel(path: Path, history: dict[str, Any], day_rows: list[dict[str, Any]], run_date: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # --- Daily Summary
    ws_sum = wb.active
    ws_sum.title = "Daily Summary"
    sum_headers = [
        "Date",
        "Runs that day",
        "Hits logged",
        "New first-seen",
        "Notes",
    ]
    for c, h in enumerate(sum_headers, 1):
        ws_sum.cell(1, c, h)
    _style_header(ws_sum, len(sum_headers))

    days = history.get("days") or {}
    for r_i, day in enumerate(sorted(days.keys(), reverse=True), 2):
        bucket = days[day]
        keys = bucket.get("notice_keys") or []
        new_keys = bucket.get("new_notice_keys") or []
        ws_sum.cell(r_i, 1, day).font = CELL_FONT
        ws_sum.cell(r_i, 2, len(bucket.get("runs") or [])).font = CELL_FONT
        ws_sum.cell(r_i, 3, len(keys)).font = CELL_FONT
        ws_sum.cell(r_i, 4, len(new_keys)).font = CELL_FONT
        note = "today's run" if day == run_date else ""
        ws_sum.cell(r_i, 5, note).font = CELL_FONT
        for c in range(1, 6):
            ws_sum.cell(r_i, c).border = THIN
            if day == run_date:
                ws_sum.cell(r_i, c).fill = DAY_FILL
    _autosize(ws_sum)
    ws_sum.freeze_panes = "A2"
    ws_sum.auto_filter.ref = f"A1:E{max(2, len(days) + 1)}"

    # --- Today's Results
    sheet_name = f"Day {run_date}"[:31]
    ws_day = wb.create_sheet(sheet_name)
    for c, (_, label) in enumerate(DETAIL_COLUMNS, 1):
        ws_day.cell(1, c, label)
    _style_header(ws_day, len(DETAIL_COLUMNS))
    _write_detail_rows(ws_day, day_rows)
    _autosize(ws_day, max_width=56)
    ws_day.freeze_panes = "A2"
    if day_rows:
        ws_day.auto_filter.ref = f"A1:{get_column_letter(len(DETAIL_COLUMNS))}{len(day_rows) + 1}"

    # --- All Notices master
    ws_all = wb.create_sheet("All Notices")
    for c, (_, label) in enumerate(DETAIL_COLUMNS, 1):
        ws_all.cell(1, c, label)
    _style_header(ws_all, len(DETAIL_COLUMNS))
    all_rows = []
    for key, rec in (history.get("notices") or {}).items():
        row = filter_notice_row(dict(rec))
        if not row:
            continue
        row["tracked_date"] = rec.get("last_seen_date") or rec.get("first_seen_date") or ""
        row["is_new"] = rec.get("first_seen_date") == run_date
        all_rows.append(row)
    all_rows.sort(key=lambda r: (r.get("last_seen_date") or "", r.get("posted_date") or ""), reverse=True)
    _write_detail_rows(ws_all, all_rows)
    _autosize(ws_all, max_width=56)
    ws_all.freeze_panes = "A2"
    if all_rows:
        ws_all.auto_filter.ref = f"A1:{get_column_letter(len(DETAIL_COLUMNS))}{len(all_rows) + 1}"

    # --- First Seen By Day (one row per notice under first_seen_date)
    ws_fs = wb.create_sheet("First Seen By Day")
    for c, (_, label) in enumerate(DETAIL_COLUMNS, 1):
        ws_fs.cell(1, c, label)
    _style_header(ws_fs, len(DETAIL_COLUMNS))
    fs_rows = []
    for rec in (history.get("notices") or {}).values():
        row = filter_notice_row(dict(rec))
        if not row:
            continue
        row["tracked_date"] = rec.get("first_seen_date") or ""
        row["is_new"] = rec.get("first_seen_date") == run_date
        fs_rows.append(row)
    fs_rows.sort(key=lambda r: (r.get("first_seen_date") or "", r.get("posted_date") or ""), reverse=True)
    _write_detail_rows(ws_fs, fs_rows)
    _autosize(ws_fs, max_width=56)
    ws_fs.freeze_panes = "A2"
    if fs_rows:
        ws_fs.auto_filter.ref = f"A1:{get_column_letter(len(DETAIL_COLUMNS))}{len(fs_rows) + 1}"

    # --- Runs log
    ws_runs = wb.create_sheet("Run Log")
    run_headers = [
        "run_id",
        "run_date",
        "posted_from",
        "posted_to",
        "term_count",
        "hit_count",
        "new_count",
        "term_sleep",
        "page_sleep",
        "errors",
    ]
    for c, h in enumerate(run_headers, 1):
        ws_runs.cell(1, c, h)
    _style_header(ws_runs, len(run_headers))
    for r_i, run in enumerate(reversed(history.get("runs") or []), 2):
        for c, h in enumerate(run_headers, 1):
            val = run.get(h, "")
            if h == "errors" and isinstance(val, list):
                val = "; ".join(val) if val else ""
            cell = ws_runs.cell(r_i, c, val)
            cell.font = CELL_FONT
            cell.border = THIN
    _autosize(ws_runs)
    ws_runs.freeze_panes = "A2"

    wb.save(path)


def write_html(path: Path, history: dict[str, Any], day_rows: list[dict[str, Any]], run_date: str, meta: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    days = history.get("days") or {}
    notices = history.get("notices") or {}

    def terms_cell(val: Any) -> str:
        if isinstance(val, list):
            return html.escape("; ".join(val))
        return html.escape(str(val or ""))

    def row_html(r: dict[str, Any]) -> str:
        new_badge = '<span class="badge new">NEW</span>' if r.get("is_new") else ""
        url = r.get("url") or ""
        title = html.escape(r.get("title") or "(no title)")
        title_html = f'<a href="{html.escape(url)}" target="_blank" rel="noopener">{title}</a>' if url else title
        return f"""
        <tr class="{'is-new' if r.get('is_new') else ''}">
          <td>{new_badge}</td>
          <td>{html.escape(str(r.get('posted_date') or ''))}</td>
          <td>{title_html}</td>
          <td>{terms_cell(r.get('matched_terms'))}</td>
          <td>{html.escape(str(r.get('type') or ''))}</td>
          <td>{html.escape(str(r.get('response_deadline') or ''))}</td>
          <td>{html.escape(str(r.get('organization') or ''))}</td>
          <td>{html.escape(str(r.get('solicitation_number') or ''))}</td>
          <td><code>{html.escape(str(r.get('notice_id') or ''))}</code></td>
        </tr>"""

    # Build sections for every day in history (newest first)
    day_sections: list[str] = []
    for day in sorted(days.keys(), reverse=True):
        bucket = days[day]
        keys = bucket.get("notice_keys") or []
        new_keys = set(bucket.get("new_notice_keys") or [])
        rows = []
        for k in keys:
            rec = notices.get(k)
            if not rec:
                continue
            row = filter_notice_row(dict(rec))
            if not row:
                continue
            row["is_new"] = k in new_keys or rec.get("first_seen_date") == day
            rows.append(row)
        rows.sort(key=lambda r: (0 if r.get("is_new") else 1, -_date_sort_key(r.get("posted_date"))))

        open_attr = " open" if day == run_date else ""
        body_rows = "".join(row_html(r) for r in rows) or '<tr><td colspan="9"><em>No hits</em></td></tr>'
        day_sections.append(
            f"""
        <details class="day"{open_attr}>
          <summary>
            <strong>{html.escape(day)}</strong>
            · {len(rows)} hit(s)
            · {len(new_keys)} new
            {" · <em>this run</em>" if day == run_date else ""}
          </summary>
          <table>
            <thead>
              <tr>
                <th></th><th>Posted</th><th>Title</th><th>Terms</th><th>Type</th>
                <th>Deadline</th><th>Organization</th><th>Solicitation</th><th>Notice ID</th>
              </tr>
            </thead>
            <tbody>{body_rows}</tbody>
          </table>
        </details>"""
        )

    errors = meta.get("errors") or []
    err_html = ""
    if errors:
        err_html = "<div class='errors'><strong>Term errors:</strong><ul>" + "".join(
            f"<li>{html.escape(e)}</li>" for e in errors
        ) + "</ul></div>"

    doc = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>SAM tracker — {html.escape(run_date)}</title>
  <style>
    :root {{
      --bg: #f6f8fb;
      --card: #ffffff;
      --ink: #1a2332;
      --muted: #5b6b7c;
      --line: #d8e0ea;
      --new: #e8f6e8;
      --accent: #1f4e79;
      --link: #0b5cab;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0; padding: 24px;
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Helvetica, Arial, sans-serif;
      color: var(--ink); background: var(--bg); line-height: 1.45;
    }}
    h1 {{ margin: 0 0 8px; font-size: 1.55rem; color: var(--accent); }}
    .meta {{ color: var(--muted); margin-bottom: 20px; }}
    .meta code {{ background: #eef2f7; padding: 1px 6px; border-radius: 4px; }}
    .card {{
      background: var(--card); border: 1px solid var(--line); border-radius: 12px;
      padding: 16px 18px; margin-bottom: 16px; box-shadow: 0 1px 2px rgba(0,0,0,.04);
    }}
    details.day {{
      background: var(--card); border: 1px solid var(--line); border-radius: 12px;
      margin-bottom: 12px; padding: 8px 12px 12px; box-shadow: 0 1px 2px rgba(0,0,0,.04);
    }}
    details.day > summary {{
      cursor: pointer; padding: 8px 4px; list-style: none; font-size: 1.02rem;
    }}
    details.day > summary::-webkit-details-marker {{ display: none; }}
    details.day > summary::before {{ content: "▸ "; color: var(--accent); }}
    details.day[open] > summary::before {{ content: "▾ "; }}
    table {{ width: 100%; border-collapse: collapse; margin-top: 8px; font-size: 0.92rem; }}
    th, td {{ border-bottom: 1px solid var(--line); padding: 8px 6px; text-align: left; vertical-align: top; }}
    th {{ background: #eef3f8; color: var(--accent); position: sticky; top: 0; }}
    tr.is-new {{ background: var(--new); }}
    a {{ color: var(--link); text-decoration: none; }}
    a:hover {{ text-decoration: underline; }}
    .badge.new {{
      display: inline-block; background: #2e7d32; color: #fff;
      font-size: 0.7rem; font-weight: 700; padding: 2px 6px; border-radius: 999px;
    }}
    code {{ font-size: 0.85em; }}
    .errors {{ background: #fff4f4; border: 1px solid #f0c2c2; border-radius: 8px; padding: 10px 12px; margin-top: 12px; }}
    .legend span {{ display: inline-block; margin-right: 14px; }}
    .swatch {{ display: inline-block; width: 12px; height: 12px; border-radius: 2px; margin-right: 4px; vertical-align: middle; }}
    .swatch.new {{ background: var(--new); border: 1px solid #9ccc9c; }}
  </style>
</head>
<body>
  <div class="card">
    <h1>SAM.gov daily tracker</h1>
    <div class="meta">
      Last run: <strong>{html.escape(run_date)}</strong>
      · Posted window <code>{html.escape(str(meta.get('posted_from') or ''))}</code>
        → <code>{html.escape(str(meta.get('posted_to') or ''))}</code>
      · Terms: {meta.get('term_count', '—')}
      · This run hits: <strong>{len(day_rows)}</strong>
      · First-seen new: <strong>{sum(1 for r in day_rows if r.get('is_new'))}</strong>
      · Source: <code>{html.escape(str(meta.get('source') or '—'))}</code>
      · Pacing: {html.escape('human' if meta.get('human') else 'fixed')}
      · History: last {html.escape(str(meta.get('history_days') or DEFAULT_HISTORY_RETENTION_DAYS))} day(s)
    </div>
    <div class="legend">
      <span><span class="swatch new"></span> New first-seen on that day</span>
      <span>Frontend path = website search (keyword). API path = title filter + key quota.</span>
    </div>
    {err_html}
  </div>

  <h2 style="color:var(--accent);font-size:1.15rem;">Results by day</h2>
  {''.join(day_sections) if day_sections else '<p><em>No history yet.</em></p>'}

  <p class="meta" style="margin-top:24px;">
    Generated {html.escape(datetime.now().strftime('%Y-%m-%d %H:%M'))} ·
    Local history: data/history.json
  </p>
</body>
</html>
"""
    path.write_text(doc, encoding="utf-8")


def parse_ptypes(raw: str | None) -> list[str] | None:
    if not raw:
        return None
    parts = [p.strip().lower() for p in raw.replace(" ", "").split(",") if p.strip()]
    bad = [p for p in parts if p not in PTYPE_LABELS]
    if bad:
        raise SystemExit(f"Unknown ptype code(s): {', '.join(bad)}. Valid: {', '.join(PTYPE_LABELS)}")
    return parts


def load_api_rotation(path: Path) -> dict[str, Any]:
    empty = {
        "next_index": 0,
        "last_run_date": None,
        "last_batch": [],
        "last_start_index": 0,
        "last_completed": 0,
        "log": [],
    }
    if not path.is_file():
        return dict(empty)
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return dict(empty)
    if not isinstance(data, dict):
        return dict(empty)
    for k, v in empty.items():
        data.setdefault(k, v)
    return data


def save_api_rotation(path: Path, state: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(state, indent=2) + "\n", encoding="utf-8")


def select_api_term_batch(
    terms: list[str],
    *,
    batch_size: int,
    state_path: Path,
    run_date: str,
    force_new: bool = False,
) -> tuple[list[str], dict[str, Any], int, str]:
    """
    Rotate through the full term list in chunks (for API daily quota).

    State machine (persisted in api_term_rotation.json):
      next_index — next term index to start a *new* day's slice (0..n-1, wraps)
      last_batch — planned full batch for last_run_date
      last_start_index — index of last_batch[0]
      last_completed — how many of last_batch finished (no double-count on re-run)

    Modes returned as mode string:
      new      — new calendar day (or force_new): take batch_size from next_index
      resume   — same day, unfinished batch: only remaining terms (save quota)
      rerun    — same day, batch already finished: re-query same batch, do not advance

    Returns (work_terms, state, work_start_index, mode).
    """
    if not terms:
        return [], load_api_rotation(state_path), 0, "empty"
    batch_size = max(1, min(int(batch_size), len(terms)))
    state = load_api_rotation(state_path)
    n = len(terms)
    term_set = set(terms)

    if not force_new and state.get("last_run_date") == run_date:
        last_batch = [t for t in (state.get("last_batch") or []) if t in term_set]
        last_completed = max(0, int(state.get("last_completed") or 0))
        last_start = int(state.get("last_start_index") or 0) % n

        if last_batch and last_completed < len(last_batch):
            # Resume: only terms not yet finished today (e.g. after 429 mid-batch)
            remainder = last_batch[last_completed:]
            work_start = (last_start + last_completed) % n
            print(
                f"API rotation: RESUME today's unfinished batch "
                f"({last_completed}/{len(last_batch)} done; {len(remainder)} left) "
                f"from index {work_start}",
                flush=True,
            )
            print("  API remaining: " + ", ".join(remainder), flush=True)
            return remainder, state, work_start, "resume"

        if last_batch and last_completed >= len(last_batch):
            print(
                f"API rotation: RE-RUN today's finished batch ({len(last_batch)} terms) "
                f"at index {last_start} — will not advance next_index",
                flush=True,
            )
            print("  API batch: " + ", ".join(last_batch), flush=True)
            return list(last_batch), state, last_start, "rerun"

    # New day (or force): slice from next_index
    start = int(state.get("next_index") or 0) % n
    batch = [terms[(start + i) % n] for i in range(batch_size)]
    days_for_full = (n + batch_size - 1) // batch_size
    print(
        f"API rotation: NEW batch {batch_size}/{n} terms "
        f"(index {start}..{(start + batch_size - 1) % n}); "
        f"full list ≈ every {days_for_full} day(s)",
        flush=True,
    )
    print("  API batch: " + ", ".join(batch), flush=True)
    return batch, state, start, "new"


def commit_api_rotation(
    state: dict[str, Any],
    *,
    state_path: Path,
    terms: list[str],
    work_terms: list[str],
    work_start: int,
    run_date: str,
    completed_count: int,
    mode: str,
    errors: list[str],
) -> None:
    """
    Persist rotation after an API phase.

    - mode=new: last_batch becomes work_terms; advance next_index by completed
    - mode=resume: keep original last_batch; last_completed += completed; advance next_index
    - mode=rerun: do not change next_index / last_completed (already finished today)
    """
    n = max(len(terms), 1)
    completed = max(0, min(int(completed_count), len(work_terms)))

    if mode == "rerun":
        # Already finished this slice today; optional re-query must not move the cursor
        state["last_run_date"] = run_date
        log = list(state.get("log") or [])
        log.append(
            {
                "date": run_date,
                "mode": mode,
                "work_terms": work_terms,
                "completed": completed,
                "next_index": state.get("next_index"),
                "errors": errors[-5:],
            }
        )
        state["log"] = log[-60:]
        save_api_rotation(state_path, state)
        print(
            f"API rotation saved: mode=rerun completed {completed}/{len(work_terms)}; "
            f"next_index unchanged={state.get('next_index')}",
            flush=True,
        )
        return

    if mode == "resume":
        planned = list(state.get("last_batch") or work_terms)
        prev_done = max(0, int(state.get("last_completed") or 0))
        # Cap so we never exceed planned batch length
        total_done = min(len(planned), prev_done + completed)
        # IMPORTANT: use .get(key, default) — not `or` — because start index 0 is valid
        if "last_start_index" in state and state["last_start_index"] is not None:
            plan_start = int(state["last_start_index"]) % n
        else:
            plan_start = work_start % n
        state["last_batch"] = planned
        state["last_start_index"] = plan_start
        state["last_completed"] = total_done
        state["next_index"] = (plan_start + total_done) % n
    else:
        # mode == "new" (or empty/fallback)
        state["last_batch"] = list(work_terms)
        state["last_start_index"] = work_start % n
        state["last_completed"] = completed
        state["next_index"] = (work_start + completed) % n

    state["last_run_date"] = run_date
    log = list(state.get("log") or [])
    log.append(
        {
            "date": run_date,
            "mode": mode,
            "work_start": work_start,
            "work_terms": work_terms,
            "completed": completed,
            "last_completed": state.get("last_completed"),
            "next_index": state.get("next_index"),
            "errors": errors[-5:],
        }
    )
    state["log"] = log[-60:]
    save_api_rotation(state_path, state)
    print(
        f"API rotation saved: mode={mode} work_completed={completed}/{len(work_terms)}; "
        f"slice_done={state.get('last_completed')}/{len(state.get('last_batch') or [])}; "
        f"next_index={state.get('next_index')}",
        flush=True,
    )


def publish_latest_sync_folder(
    *,
    xlsx_src: Path,
    html_src: Path,
    dest_dir: Path,
) -> list[Path]:
    """
    Write only our two latest deliverables into dest_dir (overwrite those names).
    Do NOT delete or touch any other files/folders already there.
    """
    dest_dir = Path(dest_dir).expanduser()
    dest_dir.mkdir(parents=True, exist_ok=True)

    written: list[Path] = []
    pairs = (
        (xlsx_src, dest_dir / "SAM-daily-latest.xlsx"),
        (html_src, dest_dir / "SAM-daily-latest.html"),
    )
    for src, dest in pairs:
        if not src.is_file():
            print(f"  skip missing source: {src}", file=sys.stderr)
            continue
        shutil.copy2(src, dest)
        written.append(dest)
        print(f"  synced → {dest}", flush=True)
    return written


def main(argv: list[str] | None = None) -> int:
    load_dotenv(ROOT / ".env")

    parser = argparse.ArgumentParser(
        description="Human-paced SAM.gov daily search → Excel + HTML day tracker"
    )
    parser.add_argument(
        "--source",
        choices=("frontend", "api", "both"),
        default="both",
        help="both = frontend (all terms) + API (rotated batch). Default both.",
    )
    parser.add_argument(
        "--days",
        type=int,
        default=DEFAULT_SEARCH_DAYS,
        help=f"Posted look-back window in days (default: {DEFAULT_SEARCH_DAYS})",
    )
    parser.add_argument(
        "--history-days",
        type=int,
        default=DEFAULT_HISTORY_RETENTION_DAYS,
        help=f"Keep this many days of history in reports (default: {DEFAULT_HISTORY_RETENTION_DAYS})",
    )
    parser.add_argument("--from", dest="date_from", help="Posted from YYYY-MM-DD")
    parser.add_argument("--to", dest="date_to", help="Posted to YYYY-MM-DD (default: today)")
    parser.add_argument("--ptype", dest="ptypes", default="", help="API only: comma ptypes o,k,r,p,s,a,...")
    parser.add_argument("--terms", type=Path, default=DEFAULT_TERMS)
    parser.add_argument("--term", action="append", default=[], help="Extra term (repeatable)")
    parser.add_argument("--limit", type=int, default=DEFAULT_LIMIT)
    parser.add_argument("--max-pages", type=int, default=DEFAULT_MAX_PAGES)
    parser.add_argument(
        "--api-batch-size",
        type=int,
        default=DEFAULT_API_BATCH_SIZE,
        help=f"API terms per day under quota (default {DEFAULT_API_BATCH_SIZE}; ~full list every 4 days)",
    )
    parser.add_argument(
        "--api-rotate-force",
        action="store_true",
        help="Advance to a new API batch even if one already ran today",
    )
    parser.add_argument(
        "--term-sleep",
        type=float,
        default=None,
        help="API fixed seconds between terms (default 5). Frontend uses human_pause().",
    )
    parser.add_argument(
        "--page-sleep",
        type=float,
        default=None,
        help="API fixed seconds between pages.",
    )
    parser.add_argument(
        "--jitter",
        type=float,
        default=DEFAULT_JITTER,
        help="API random extra 0..jitter seconds",
    )
    parser.add_argument(
        "--robot",
        action="store_true",
        help="Disable human-like variable pauses on frontend (not recommended)",
    )
    parser.add_argument(
        "--active-only",
        action="store_true",
        default=True,
        help="Frontend: prefer active notices (default on)",
    )
    parser.add_argument(
        "--include-inactive",
        action="store_true",
        help="Frontend: include inactive/archived in SGS query",
    )
    parser.add_argument("--base-url", default=None, help="Override search base URL (single-source only)")
    parser.add_argument("--api-key", default=os.environ.get("SAM_API_KEY", ""))
    parser.add_argument("--history", type=Path, default=DEFAULT_HISTORY)
    parser.add_argument(
        "--rotation-state",
        type=Path,
        default=DEFAULT_API_ROTATION,
        help="JSON file that remembers API term position across months of daily runs",
    )
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--report-only", action="store_true", help="Rebuild Excel/HTML from history only")
    parser.add_argument("--no-history-write", action="store_true", help="Do not update history.json")
    parser.add_argument(
        "--upload-drive",
        action="store_true",
        help="Upload via service account API (optional; prefer local Latest folder)",
    )
    parser.add_argument(
        "--no-upload-drive",
        action="store_true",
        help="Skip service-account Drive API upload even if SAM_UPLOAD_DRIVE=1",
    )
    parser.add_argument(
        "--latest-dir",
        type=Path,
        default=None,
        help="Local folder for latest copies (or set SAM_LATEST_SYNC_DIR)",
    )
    parser.add_argument(
        "--no-latest-sync",
        action="store_true",
        help="Do not copy latest xlsx/html into the Latest sync folder",
    )
    args = parser.parse_args(argv)

    source = args.source
    do_frontend = source in ("frontend", "both")
    do_api = source in ("api", "both")
    human = not args.robot
    if args.term_sleep is None:
        args.term_sleep = DEFAULT_TERM_SLEEP_API
    if args.page_sleep is None:
        args.page_sleep = DEFAULT_PAGE_SLEEP_API
    frontend_url = os.environ.get("SAM_FRONTEND_URL") or DEFAULT_FRONTEND_URL
    api_url = os.environ.get("SAM_BASE_URL") or DEFAULT_BASE_URL
    if args.base_url:
        if do_api and not do_frontend:
            api_url = args.base_url
        elif do_frontend and not do_api:
            frontend_url = args.base_url

    posted_to = parse_date(args.date_to) if args.date_to else date.today()
    if args.date_from:
        posted_from = parse_date(args.date_from)
    else:
        if args.days < 1:
            raise SystemExit("--days must be >= 1")
        posted_from = posted_to - timedelta(days=args.days - 1)

    if posted_from > posted_to:
        raise SystemExit("posted from date is after posted to date")
    if (posted_to - posted_from).days > 366:
        raise SystemExit("Date window max ~1 year")

    terms = load_terms(args.terms)
    for t in args.term:
        t = t.strip()
        if t and t.casefold() not in {x.casefold() for x in terms}:
            terms.append(t)

    ptypes = parse_ptypes(args.ptypes or None)
    api_key = normalize_api_key(args.api_key or "") or normalize_api_key(load_api_key_from_files() or "")
    active_only = not args.include_inactive

    run_date = date.today().isoformat()
    run_id = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    days = (posted_to - posted_from).days + 1

    api_batch: list[str] = []
    api_rot_state: dict[str, Any] = {}
    api_start = 0
    api_mode = "empty"
    if do_api:
        api_batch, api_rot_state, api_start, api_mode = select_api_term_batch(
            terms,
            batch_size=args.api_batch_size,
            state_path=args.rotation_state,
            run_date=run_date,
            force_new=args.api_rotate_force,
        )

    fe_est = len(terms) * 55.0 if do_frontend and human else 0.0
    api_est = len(api_batch) * (args.term_sleep + args.jitter / 2) if do_api else 0.0
    est_time = fe_est + api_est

    print(f"Source: {source}" + (" (frontend human + API rotate)" if source == "both" else ""))
    print(f"SAM search window: {posted_from.isoformat()} → {posted_to.isoformat()} ({days} day(s))")
    if do_frontend:
        print(f"Frontend: ALL {len(terms)} terms · human pacing · {frontend_url}")
    if do_api:
        print(
            f"API: {len(api_batch)}/{len(terms)} terms today (batch size {args.api_batch_size}) · "
            f"sleep {args.term_sleep}s · {api_url}"
        )
        print(f"API key: {'set' if api_key else 'MISSING'} (len={len(api_key) if api_key else 0})")
        print(
            "API rotation: does NOT reset each day — position saved in "
            f"{args.rotation_state.name}; wraps after full list (months OK)."
        )
    print(f"Est. rough wall time ≥ {est_time/60:.0f} min (plus network)")

    if args.dry_run:
        if do_frontend:
            print("Frontend terms (all):")
            for t in terms:
                print(f"  - {t}")
        if do_api:
            print(f"API work today (mode={api_mode}):")
            for t in api_batch:
                print(f"  - {t}")
            rot = load_api_rotation(args.rotation_state)
            print(f"  work_start={api_start}  saved next_index={rot.get('next_index', 0)}")
            print(f"  (dry-run does not advance rotation)")
        return 0

    history = load_history(args.history)
    errors: list[str] = []
    merged: list[Hit] = []
    day_rows: list[dict[str, Any]] = []

    if args.report_only:
        print("Report-only mode: rebuilding Excel/HTML from history.json")
        prune_history(history, keep_days=max(1, args.history_days))
        day_bucket = (history.get("days") or {}).get(run_date) or {}
        keys = day_bucket.get("notice_keys") or []
        new_keys = set(day_bucket.get("new_notice_keys") or [])
        for k in keys:
            rec = (history.get("notices") or {}).get(k)
            if not rec:
                continue
            row = filter_notice_row(dict(rec))
            if not row:
                continue
            row["tracked_date"] = run_date
            row["is_new"] = k in new_keys
            day_rows.append(row)
        print(
            f"After phrase filter: {len(day_rows)} row(s) for {run_date}; "
            f"history days kept: {sorted((history.get('days') or {}).keys())}",
            flush=True,
        )
    else:
        if do_api and not api_key:
            print(
                "\nNo SAM_API_KEY set (required for API portion).\n"
                "Put key in .env or api.txt, or use --source frontend only.\n",
                file=sys.stderr,
            )
            if source == "api":
                return 2
            print("Continuing with frontend only.", flush=True)
            do_api = False

        all_hits: list[Hit] = []
        run_started = time.monotonic()

        def run_phase(
            *,
            phase_name: str,
            phase_terms: list[str],
            use_frontend: bool,
        ) -> int:
            """Return number of terms completed without quota stop."""
            completed = 0
            n = len(phase_terms)
            if n == 0:
                return 0
            print(
                f"\n=== {phase_name}: {n} term(s) ===",
                flush=True,
            )
            for i, term in enumerate(phase_terms, 1):
                print_progress(
                    done=i - 1,
                    total=n,
                    term=term,
                    hits_this=None,
                    total_hits=len(all_hits),
                    errors=len(errors),
                    started=run_started,
                    phase="fetch",
                )
                q_disp = search_query_for_term(term)
                label = f"[{phase_name} {i}/{n}] {term!r}" + (
                    f" as {q_disp}" if q_disp != term else ""
                )
                print(("\n" if sys.stdout.isatty() else "") + f"{label} ...", flush=True)
                try:
                    if use_frontend:
                        hits = search_term_frontend(
                            base_url=frontend_url,
                            term=term,
                            posted_from=posted_from,
                            posted_to=posted_to,
                            limit=min(max(args.limit, 1), 100),
                            max_pages=max(args.max_pages, 1),
                            active_only=active_only,
                            human=human,
                        )
                    else:
                        hits = search_term_api(
                            base_url=api_url,
                            api_key=api_key,
                            term=term,
                            posted_from=posted_from,
                            posted_to=posted_to,
                            ptypes=ptypes,
                            limit=min(max(args.limit, 1), 1000),
                            max_pages=max(args.max_pages, 1),
                            page_sleep=args.page_sleep,
                            jitter=args.jitter,
                        )
                    all_hits.extend(hits)
                    completed += 1
                    print(f"    → {len(hits)} in window", flush=True)
                    print_progress(
                        done=i,
                        total=n,
                        term=term,
                        hits_this=len(hits),
                        total_hits=len(all_hits),
                        errors=len(errors),
                        started=run_started,
                        phase="done",
                    )
                    if sys.stdout.isatty():
                        print(flush=True)
                except Exception as e:  # noqa: BLE001
                    print(f"    ERROR: {e}", file=sys.stderr, flush=True)
                    errors.append(f"{phase_name} term={term!r}: {e}")
                    print_progress(
                        done=i,
                        total=n,
                        term=term,
                        hits_this=0,
                        total_hits=len(all_hits),
                        errors=len(errors),
                        started=run_started,
                        phase="ERROR",
                    )
                    if sys.stdout.isatty():
                        print(flush=True)
                    if "429" in str(e) or "throttl" in str(e).lower() or "quota" in str(e).lower():
                        print(
                            "Stopping this phase early: rate limit / quota. "
                            "Will write what we have.",
                            flush=True,
                        )
                        break
                    # Non-quota error: still count as "done" so rotation can advance
                    completed += 1

                if i < n:
                    if use_frontend and human:
                        human_pause(kind="between_terms", label=f"after {term}")
                    else:
                        paced_sleep(args.term_sleep, args.jitter, label=f"after {term}")
            return completed

        # API first (small batch, quota-sensitive), then frontend full list
        api_completed = 0
        if do_api and api_batch:
            api_completed = run_phase(
                phase_name="API",
                phase_terms=api_batch,
                use_frontend=False,
            )
            # Reload state in case file changed; use the select-time state as base
            commit_api_rotation(
                api_rot_state,
                state_path=args.rotation_state,
                terms=terms,
                work_terms=api_batch,
                work_start=api_start,
                run_date=run_date,
                completed_count=api_completed,
                mode=api_mode,
                errors=errors,
            )
        if do_frontend:
            run_phase(
                phase_name="Frontend",
                phase_terms=terms,
                use_frontend=True,
            )

        if sys.stdout.isatty():
            print(flush=True)
        merged = merge_hits(all_hits)
        print(f"\nUnique notices this run: {len(merged)}", flush=True)

        meta_partial = {
            "posted_from": posted_from.isoformat(),
            "posted_to": posted_to.isoformat(),
            "term_count": len(terms),
            "api_batch": api_batch,
            "api_mode": api_mode,
            "api_completed": api_completed if do_api else 0,
            "errors": errors,
            "term_sleep": args.term_sleep,
            "page_sleep": args.page_sleep,
            "source": source,
            "human": human,
        }
        day_rows, newly = merge_into_history(
            history,
            merged,
            run_date=run_date,
            run_id=run_id,
            meta=meta_partial,
        )
        print(f"First-seen new: {len(newly)}")
        prune_history(history, keep_days=max(1, args.history_days))
        print(
            f"History retention: last {args.history_days} day(s) "
            f"(cutoff {history.get('retention_cutoff')}); "
            f"{len(history.get('days') or {})} day bucket(s), "
            f"{len(history.get('notices') or {})} notice(s)",
            flush=True,
        )
        if not args.no_history_write:
            save_history(args.history, history, keep_days=max(1, args.history_days))
            print(f"History: {args.history}")

    meta = {
        "run_date": run_date,
        "posted_from": posted_from.isoformat(),
        "posted_to": posted_to.isoformat(),
        "term_count": len(terms),
        "errors": errors,
        "term_sleep": args.term_sleep,
        "page_sleep": args.page_sleep,
        "source": source,
        "human": human,
        "api_batch": api_batch,
        "api_mode": api_mode,
        "history_days": max(1, args.history_days),
    }

    # Primary: project root, easy to spot by date (fresh run each day)
    # e.g. SAM-daily-2026-07-15.xlsx
    xlsx_root = ROOT / f"SAM-daily-{run_date}.xlsx"
    html_root = ROOT / f"SAM-daily-{run_date}.html"
    xlsx_latest_root = ROOT / "SAM-daily-latest.xlsx"
    html_latest_root = ROOT / "SAM-daily-latest.html"
    # Archive copies under reports/
    xlsx_archive = DEFAULT_REPORTS / f"sam-tracker-{run_id}.xlsx"
    html_archive = DEFAULT_REPORTS / f"sam-tracker-{run_id}.html"

    write_excel(xlsx_root, history, day_rows, run_date)
    write_excel(xlsx_latest_root, history, day_rows, run_date)
    write_excel(xlsx_archive, history, day_rows, run_date)
    write_html(html_root, history, day_rows, run_date, meta)
    write_html(html_latest_root, history, day_rows, run_date, meta)
    write_html(html_archive, history, day_rows, run_date, meta)

    print(f"Excel (root):  {xlsx_root}")
    print(f"HTML  (root):  {html_root}")
    print(f"Latest:        {xlsx_latest_root}")
    print(f"               {html_latest_root}")
    print(f"Archive:       {xlsx_archive}")

    for h in (merged or [])[:20]:
        terms_s = ",".join(sorted(h.matched_terms, key=str.casefold))
        print(f"- [{h.posted_date}] {h.title[:90]}  ({terms_s})")
    if len(merged) > 20:
        print(f"... and {len(merged) - 20} more")

    # Always write our two latest files into Drive-for-desktop Latest folder
    # (unless --no-latest-sync). Overwrite only those names; leave other files alone.
    if not args.no_latest_sync:
        env_latest = (os.environ.get("SAM_LATEST_SYNC_DIR") or "").strip()
        latest_dir: Path | None = None
        if args.latest_dir:
            latest_dir = Path(args.latest_dir).expanduser()
        elif env_latest:
            latest_dir = Path(env_latest).expanduser()
        elif DEFAULT_LATEST_SYNC_DIR is not None:
            latest_dir = Path(DEFAULT_LATEST_SYNC_DIR).expanduser()

        if latest_dir is None:
            print(
                "\nLatest folder sync skipped (set SAM_LATEST_SYNC_DIR or --latest-dir).",
                flush=True,
            )
        else:
            print(f"\nPublishing latest files to: {latest_dir}", flush=True)
            try:
                publish_latest_sync_folder(
                    xlsx_src=xlsx_latest_root,
                    html_src=html_latest_root,
                    dest_dir=latest_dir,
                )
            except Exception as e:  # noqa: BLE001
                print(f"Latest folder sync failed: {e}", file=sys.stderr)

    # Optional Google Drive API upload (service account) — off unless requested
    env_upload = (os.environ.get("SAM_UPLOAD_DRIVE") or "").strip().lower() in (
        "1",
        "true",
        "yes",
        "on",
    )
    do_upload = (args.upload_drive or env_upload) and not args.no_upload_drive

    if do_upload:
        print("\nUploading via Drive API (service account)...", flush=True)
        try:
            sys.path.insert(0, str(Path(__file__).resolve().parent))
            from drive_upload import upload_paths  # type: ignore

            upload_paths([xlsx_latest_root, html_latest_root])
        except SystemExit as e:
            print(f"Drive API upload skipped/failed: {e}", file=sys.stderr)
        except Exception as e:  # noqa: BLE001
            print(f"Drive API upload failed: {e}", file=sys.stderr)

    return 1 if errors and not merged else 0


if __name__ == "__main__":
    raise SystemExit(main())
